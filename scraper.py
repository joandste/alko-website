from openpyxl import load_workbook
#import requests
import os
from operator import itemgetter
from jinja2 import Environment, FileSystemLoader, select_autoescape

#url = 'https://www.alko.fi/INTERSHOP/static/WFS/Alko-OnlineShop-Site/-/Alko-OnlineShop/fi_FI/Alkon%20Hinnasto%20Tekstitiedostona/alkon-hinnasto-tekstitiedostona.xlsx'
#document = requests.get(url)
#with open('alkon-hinnasto-tekstitiedostona.xlsx', 'wb') as file:
#    file.write(document.content)

wb = load_workbook(filename ='alkon-hinnasto-tekstitiedostona.xlsx')
ws = wb.active

package_path = os.path.dirname(os.path.abspath(__file__))
templates_dir = os.path.join(package_path, "templates")
env = Environment(
    loader=FileSystemLoader(templates_dir),
    autoescape=select_autoescape(["html", "xml", "j2"])
)
template = env.get_template("template.j2")

clean_rows = []

f = open('index.html', 'w', encoding='utf-8')

for row in ws.iter_rows(min_row=5, values_only=True):
    if row[21] == None or float(row[21]) == 0:
        continue
    # 0-5, 8, 21
    clean_row = dict(id=row[0], name=row[1], liter=row[3], cost=row[4], type=row[8], alkohol=row[21], alkohol_cost_per_liter=round(float(row[5]) / float(row[21]) * 100, 2))
    clean_rows.append(clean_row)

sorted_clean_rows = sorted(clean_rows, key=itemgetter('alkohol_cost_per_liter'))

f.write(template.render(rows=sorted_clean_rows))
f.close()